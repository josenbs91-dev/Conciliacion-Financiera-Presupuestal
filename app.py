import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.title("Conciliación Financiera Presupuestal - Procesos 1, 2 y 3")

uploaded_file = st.file_uploader("📂 Sube tu archivo Excel", type=["xlsx"])

if uploaded_file:
    # Leer archivo original
    df = pd.read_excel(uploaded_file, dtype=object)

    # Normalizar numéricos
    for col in ["haber", "debe"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    st.subheader("👀 Vista previa datos originales")
    st.dataframe(df.head())

    # --------------------------
    # PROCESO 1
    # --------------------------
    proceso1 = df[df["mayor"].astype(str).str.startswith(("5", "4"))].copy()
    proceso1["mayor_subcta"] = proceso1["mayor"].astype(str) + "-" + proceso1["sub_cta"].astype(str)
    proceso1 = proceso1[["mayor_subcta", "clasificador"]].drop_duplicates()

    # --------------------------
    # PROCESO 2 (Filtros contables)
    # --------------------------
    filtro1 = df[
        (df["tipo_ctb"].astype(str) == "1") &
        (df["haber"] != 0) &
        (
            ((df["ciclo"] == "G") & (df["fase"] == "D")) |
            ((df["ciclo"] == "I") & (df["fase"] == "D"))
        )
    ].copy()
    filtro1["saldo"] = filtro1["haber"]

    filtro2 = df[
        (df["tipo_ctb"].astype(str) == "2") &
        (df["debe"] != 0) &
        (
            ((df["ciclo"] == "G") & (df["fase"] == "D")) |
            ((df["ciclo"] == "I") & (df["fase"] == "R"))
        )
    ].copy()
    filtro2["saldo"] = filtro2["debe"]

    filtro3 = df[
        (df["ciclo"] == "C") & (df["fase"] == "C") &
        (
            df["mayor"].astype(str).str.startswith("5") |
            df["mayor"].astype(str).str.startswith("4") |
            df["mayor"].astype(str).str.startswith("8501") |
            df["mayor"].astype(str).str.startswith("8601")
        )
    ].copy()
    filtro3["saldo"] = pd.to_numeric(filtro3["haber"], errors="coerce").fillna(0) - \
                       pd.to_numeric(filtro3["debe"], errors="coerce").fillna(0)

    filtrado = pd.concat([filtro1, filtro2, filtro3])
    filtrado["codigo_unido"] = (
        filtrado["mayor"].astype(str) + "-" +
        filtrado["sub_cta"].astype(str) + "-" +
        filtrado["clasificador"].astype(str)
    )

    columnas_finales = [
        "codigo_unido", "nro_not_exp", "desc_documento",
        "nro_doc", "Fecha Contable", "desc_proveedor", "saldo"
    ]
    proceso2 = filtrado[columnas_finales]

    st.subheader("📊 Proceso 1")
    st.dataframe(proceso1)

    st.subheader("📊 Proceso 2")
    st.dataframe(proceso2)

    # --------------------------
    # PROCESO 3 (Conciliación)
    # --------------------------
    conciliacion_tables = []
    for _, row in proceso1.iterrows():
        mayor_subcta = str(row["mayor_subcta"])
        clasificador = str(row["clasificador"])

        tabla = proceso2[
            proceso2["codigo_unido"].str.contains(mayor_subcta, na=False) |
            proceso2["codigo_unido"].str.contains(clasificador, na=False)
        ].copy()

        if not tabla.empty:
            conciliacion_tables.append((mayor_subcta, clasificador, tabla))

    st.subheader("📊 Proceso 3 (Conciliación)")
    st.write(f"Se generaron {len(conciliacion_tables)} tablas de conciliación")

    # --------------------------
    # EXPORTAR A EXCEL
    # --------------------------
    output_file = "resultado_procesos.xlsx"

    # Guardar hojas Proceso 1 y 2
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Original", index=False)
        proceso1.to_excel(writer, sheet_name="Proceso 1", index=False)
        proceso2.to_excel(writer, sheet_name="Proceso 2", index=False)

    # Abrir libro y agregar hoja Conciliacion manualmente
    wb = load_workbook(output_file)
    ws = wb.create_sheet("Conciliacion")

    row_start = 1
    for mayor_subcta, clasificador, tabla in conciliacion_tables:
        # Escribir título
        ws.cell(row=row_start, column=1, value=f"Mayor-Subcta: {mayor_subcta} | Clasificador: {clasificador}")
        row_start += 1

        # Escribir tabla
        for r in dataframe_to_rows(tabla, index=False, header=True):
            ws.append(r)
        row_start = ws.max_row + 6  # dejar 5 filas en blanco

    wb.save(output_file)

    # Botón de descarga
    with open(output_file, "rb") as f:
        st.download_button(
            "⬇️ Descargar Excel con Procesos 1, 2 y 3",
            f,
            file_name="resultado_procesos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
